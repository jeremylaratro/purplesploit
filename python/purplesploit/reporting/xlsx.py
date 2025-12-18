"""
Excel Report Generator for PurpleSploit.

Generates Excel reports with multiple worksheets and pivot table support.
"""

from pathlib import Path
from typing import Optional, Dict, Any, List
from datetime import datetime

from .models import ReportData, Finding, Severity


class XLSXReportGenerator:
    """Generates Excel reports from ReportData"""

    def __init__(self):
        """Initialize Excel generator"""
        self._openpyxl_available = self._check_openpyxl()

    def _check_openpyxl(self) -> bool:
        """Check if openpyxl is available"""
        try:
            import openpyxl
            return True
        except ImportError:
            return False

    def generate(
        self,
        report_data: ReportData,
        output_path: Path,
        **kwargs
    ) -> str:
        """
        Generate Excel report.

        Args:
            report_data: Report data to render
            output_path: Output file path
            **kwargs: Additional options

        Returns:
            Path to generated report

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not self._openpyxl_available:
            raise ImportError(
                "openpyxl is required for Excel generation. "
                "Install with: pip install openpyxl"
            )

        from openpyxl import Workbook
        from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import PieChart, Reference

        wb = Workbook()

        # Create worksheets
        self._create_summary_sheet(wb, report_data)
        self._create_findings_sheet(wb, report_data)
        self._create_targets_sheet(wb, report_data)
        self._create_services_sheet(wb, report_data)

        # Remove default sheet if still exists
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Save workbook
        output_path = Path(output_path)
        wb.save(output_path)

        return str(output_path)

    def _create_summary_sheet(self, wb, report_data: ReportData):
        """Create summary/dashboard sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import PieChart, Reference

        ws = wb.create_sheet("Summary", 0)

        # Styles
        title_font = Font(size=18, bold=True)
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        # Title
        ws['A1'] = report_data.config.title
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        # Report metadata
        row = 3
        metadata = [
            ("Client", report_data.config.client_name),
            ("Assessor", report_data.config.assessor_name),
            ("Report Date", report_data.config.report_date.strftime("%Y-%m-%d") if report_data.config.report_date else ""),
            ("Assessment Type", report_data.config.assessment_type),
        ]

        for label, value in metadata:
            if value:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1

        # Severity summary table
        row += 2
        ws[f'A{row}'] = "Findings by Severity"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        severity_headers = ["Severity", "Count"]
        for col, header in enumerate(severity_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row += 1
        severity_colors = {
            "critical": "7B241C",
            "high": "C0392B",
            "medium": "E67E22",
            "low": "F1C40F",
            "info": "3498DB",
        }

        chart_start_row = row
        for severity in Severity:
            count = report_data.severity_counts.get(severity.value, 0)
            ws.cell(row=row, column=1, value=severity.value.upper())
            ws.cell(row=row, column=2, value=count)

            # Color the severity cell
            color = severity_colors.get(severity.value, "95A5A6")
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            ws.cell(row=row, column=1).font = Font(color="FFFFFF", bold=True)

            row += 1

        # Total row
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=report_data.total_findings)
        ws.cell(row=row, column=2).font = Font(bold=True)

        # Create pie chart
        if report_data.total_findings > 0:
            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_start_row + 4)
            data = Reference(ws, min_col=2, min_row=chart_start_row - 1, max_row=chart_start_row + 4)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Severity Distribution"
            pie.width = 12
            pie.height = 8
            ws.add_chart(pie, "D5")

        # Findings by target
        row += 3
        ws[f'A{row}'] = "Findings by Target"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        headers = ["Target", "Findings"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row += 1
        for target, findings in report_data.findings_by_target.items():
            ws.cell(row=row, column=1, value=target)
            ws.cell(row=row, column=2, value=len(findings))
            row += 1

        # Auto-size columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_findings_sheet(self, wb, report_data: ReportData):
        """Create detailed findings sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Findings")

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        # Headers
        headers = [
            "ID", "Title", "Severity", "CVSS", "Target", "Port", "Service",
            "Description", "Impact", "Remediation", "CVEs", "Status", "Module"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Severity colors for cells
        severity_fills = {
            Severity.CRITICAL: PatternFill(start_color="7B241C", end_color="7B241C", fill_type="solid"),
            Severity.HIGH: PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid"),
            Severity.MEDIUM: PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid"),
            Severity.LOW: PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid"),
            Severity.INFO: PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid"),
        }

        # Data rows
        row = 2
        for finding in report_data.findings:
            ws.cell(row=row, column=1, value=finding.id)
            ws.cell(row=row, column=2, value=finding.title)

            # Severity with color
            sev_cell = ws.cell(row=row, column=3, value=finding.severity.value.upper())
            sev_cell.fill = severity_fills.get(finding.severity)
            sev_cell.font = Font(color="FFFFFF", bold=True)

            ws.cell(row=row, column=4, value=finding.cvss_score or "")
            ws.cell(row=row, column=5, value=finding.target)
            ws.cell(row=row, column=6, value=finding.port or "")
            ws.cell(row=row, column=7, value=finding.service or "")

            desc_cell = ws.cell(row=row, column=8, value=finding.description)
            desc_cell.alignment = wrap_alignment

            impact_cell = ws.cell(row=row, column=9, value=finding.impact)
            impact_cell.alignment = wrap_alignment

            rem_cell = ws.cell(row=row, column=10, value=finding.remediation)
            rem_cell.alignment = wrap_alignment

            ws.cell(row=row, column=11, value=", ".join(finding.cve_ids) if finding.cve_ids else "")
            ws.cell(row=row, column=12, value=finding.status.value)
            ws.cell(row=row, column=13, value=finding.module_name or "")

            row += 1

        # Auto-size columns
        column_widths = [8, 40, 12, 8, 20, 8, 15, 50, 40, 40, 20, 12, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add autofilter
        ws.auto_filter.ref = f"A1:M{row - 1}"

    def _create_targets_sheet(self, wb, report_data: ReportData):
        """Create targets sheet"""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Targets")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        headers = ["Name", "IP/URL", "Description", "Findings Count"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        for target in report_data.targets:
            target_id = target.get("ip") or target.get("url", "")
            findings_count = len(report_data.findings_by_target.get(target_id, []))

            ws.cell(row=row, column=1, value=target.get("name", ""))
            ws.cell(row=row, column=2, value=target_id)
            ws.cell(row=row, column=3, value=target.get("description", ""))
            ws.cell(row=row, column=4, value=findings_count)
            row += 1

        # Auto-size columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

        ws.freeze_panes = 'A2'

    def _create_services_sheet(self, wb, report_data: ReportData):
        """Create services sheet"""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Services")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        headers = ["Target", "Port", "Service", "Version"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        for service in report_data.services:
            ws.cell(row=row, column=1, value=service.get("target", ""))
            ws.cell(row=row, column=2, value=service.get("port", ""))
            ws.cell(row=row, column=3, value=service.get("service", ""))
            ws.cell(row=row, column=4, value=service.get("version", ""))
            row += 1

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

        ws.freeze_panes = 'A2'
        if row > 2:
            ws.auto_filter.ref = f"A1:D{row - 1}"


def get_column_letter(col_idx):
    """Get Excel column letter from index (1-indexed)"""
    from openpyxl.utils import get_column_letter as _get_column_letter
    return _get_column_letter(col_idx)
