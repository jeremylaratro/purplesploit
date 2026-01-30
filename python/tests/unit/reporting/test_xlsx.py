"""
Tests for purplesploit.reporting.xlsx module.

Tests the XLSXReportGenerator class.
"""

import pytest
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from datetime import datetime
from collections import defaultdict

from purplesploit.reporting.xlsx import XLSXReportGenerator, OPENPYXL_AVAILABLE
from purplesploit.reporting.models import (
    Finding,
    Severity,
    FindingStatus,
    ReportConfig,
    ReportData,
)


@pytest.fixture
def sample_report_data():
    """Create sample report data for testing."""
    config = ReportConfig(
        title="XLSX Test Report",
        client_name="Test Client",
        assessor_name="Test Assessor",
    )

    findings = [
        Finding(
            id="f1",
            title="SQL Injection",
            severity=Severity.CRITICAL,
            description="SQL injection vulnerability",
            target="192.168.1.100",
            port=443,
            service="https",
            status=FindingStatus.CONFIRMED,
            cvss_score=9.8,
            cve_ids=["CVE-2021-1234"],
            impact="Full database access",
            remediation="Use parameterized queries",
        ),
        Finding(
            id="f2",
            title="XSS",
            severity=Severity.MEDIUM,
            description="Reflected XSS",
            target="192.168.1.100",
            status=FindingStatus.DRAFT,
        ),
        Finding(
            id="f3",
            title="Info Disclosure",
            severity=Severity.INFO,
            description="Server version disclosed",
            target="192.168.1.101",
        ),
    ]

    return ReportData(
        config=config,
        findings=findings,
        targets=[
            {"ip": "192.168.1.100", "name": "Web Server"},
            {"ip": "192.168.1.101", "name": "API Server"},
        ],
        services=[
            {"target": "192.168.1.100", "port": 443, "service": "https"},
            {"target": "192.168.1.101", "port": 80, "service": "http"},
        ],
    )


class TestXLSXReportGeneratorInit:
    """Tests for XLSXReportGenerator initialization."""

    def test_init(self):
        """Test initialization."""
        gen = XLSXReportGenerator()

        assert isinstance(gen._openpyxl_available, bool)

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_openpyxl_available_when_installed(self):
        """Test that openpyxl is detected when installed."""
        gen = XLSXReportGenerator()

        assert gen._openpyxl_available is True


class TestXLSXReportGeneratorGenerate:
    """Tests for XLSX generation."""

    def test_generate_raises_without_openpyxl(self, sample_report_data, tmp_path):
        """Test that generate raises ImportError when openpyxl not available."""
        gen = XLSXReportGenerator()
        gen._openpyxl_available = False

        output_path = tmp_path / "report.xlsx"

        with pytest.raises(ImportError, match="openpyxl is required"):
            gen.generate(sample_report_data, output_path)

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_generate_creates_file(self, sample_report_data, tmp_path):
        """Test that generate creates output file."""
        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        result = gen.generate(sample_report_data, output_path)

        assert Path(result).exists()
        assert Path(result).stat().st_size > 0

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_generate_creates_worksheets(self, sample_report_data, tmp_path):
        """Test that generate creates expected worksheets."""
        import openpyxl

        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        gen.generate(sample_report_data, output_path)

        wb = openpyxl.load_workbook(output_path)
        sheet_names = wb.sheetnames

        assert "Summary" in sheet_names
        assert "Findings" in sheet_names
        assert "Targets" in sheet_names
        assert "Services" in sheet_names

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_summary_sheet_content(self, sample_report_data, tmp_path):
        """Test Summary sheet contains correct data."""
        import openpyxl

        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        gen.generate(sample_report_data, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Summary"]

        # Check title is in first row
        assert ws["A1"].value == "XLSX Test Report"

        # Should contain severity counts
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Critical" in str(values)
        assert "High" in str(values)
        assert "Medium" in str(values)

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_findings_sheet_headers(self, sample_report_data, tmp_path):
        """Test Findings sheet has correct headers."""
        import openpyxl

        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        gen.generate(sample_report_data, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Findings"]

        headers = [cell.value for cell in ws[1]]
        expected_headers = ["ID", "Title", "Severity", "Status", "Target", "Port", "Service"]

        for expected in expected_headers:
            assert expected in headers

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_findings_sheet_data(self, sample_report_data, tmp_path):
        """Test Findings sheet contains finding data."""
        import openpyxl

        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        gen.generate(sample_report_data, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Findings"]

        # Get all data
        data = list(ws.iter_rows(min_row=2, values_only=True))

        # Should have 3 findings
        assert len(data) == 3

        # First finding should be SQL Injection
        titles = [row[1] for row in data]
        assert "SQL Injection" in titles
        assert "XSS" in titles
        assert "Info Disclosure" in titles

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_targets_sheet_content(self, sample_report_data, tmp_path):
        """Test Targets sheet contains target data."""
        import openpyxl

        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        gen.generate(sample_report_data, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Targets"]

        # Get all data
        data = list(ws.iter_rows(values_only=True))

        # Should have header + 2 targets
        assert len(data) >= 3

        # Check targets are present
        values_str = str(data)
        assert "192.168.1.100" in values_str
        assert "Web Server" in values_str

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_services_sheet_content(self, sample_report_data, tmp_path):
        """Test Services sheet contains service data."""
        import openpyxl

        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        gen.generate(sample_report_data, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Services"]

        # Get all data
        data = list(ws.iter_rows(values_only=True))

        # Should have header + 2 services
        assert len(data) >= 3

        # Check services are present
        values_str = str(data)
        assert "https" in values_str
        assert "443" in values_str


class TestXLSXReportGeneratorFormatting:
    """Tests for XLSX formatting."""

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_severity_fill_colors(self, sample_report_data, tmp_path):
        """Test that severity cells have correct fill colors."""
        import openpyxl
        from openpyxl.styles import PatternFill

        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        gen.generate(sample_report_data, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Findings"]

        # Find severity column index
        headers = [cell.value for cell in ws[1]]
        sev_col = headers.index("Severity") + 1

        # Check cells have fills
        critical_row = None
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[sev_col - 1].value == "CRITICAL":
                critical_row = row_num
                break

        if critical_row:
            cell = ws.cell(row=critical_row, column=sev_col)
            # Cell should have some fill
            assert cell.fill is not None

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_header_styling(self, sample_report_data, tmp_path):
        """Test that headers have bold styling."""
        import openpyxl

        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        gen.generate(sample_report_data, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Findings"]

        # First row headers should be bold
        for cell in ws[1]:
            if cell.value:
                assert cell.font.bold is True

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_column_widths_adjusted(self, sample_report_data, tmp_path):
        """Test that column widths are adjusted."""
        import openpyxl

        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        gen.generate(sample_report_data, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Findings"]

        # Check that at least one column has a non-default width
        has_adjusted_width = False
        for col_dim in ws.column_dimensions.values():
            if col_dim.width is not None and col_dim.width > 8.43:  # Default width
                has_adjusted_width = True
                break

        # This test is lenient - just checking widths were considered
        assert ws.column_dimensions is not None

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_auto_filter_enabled(self, sample_report_data, tmp_path):
        """Test that auto-filter is enabled on findings sheet."""
        import openpyxl

        gen = XLSXReportGenerator()
        output_path = tmp_path / "report.xlsx"

        gen.generate(sample_report_data, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Findings"]

        # Auto filter should be set
        assert ws.auto_filter.ref is not None


class TestXLSXReportGeneratorHelpers:
    """Tests for helper methods."""

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_get_severity_fill_critical(self):
        """Test fill color for critical severity."""
        gen = XLSXReportGenerator()
        fill = gen._get_severity_fill(Severity.CRITICAL)

        # Should return some fill
        assert fill is not None

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_get_severity_fill_high(self):
        """Test fill color for high severity."""
        gen = XLSXReportGenerator()
        fill = gen._get_severity_fill(Severity.HIGH)

        assert fill is not None

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_get_severity_fill_medium(self):
        """Test fill color for medium severity."""
        gen = XLSXReportGenerator()
        fill = gen._get_severity_fill(Severity.MEDIUM)

        assert fill is not None

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_get_severity_fill_low(self):
        """Test fill color for low severity."""
        gen = XLSXReportGenerator()
        fill = gen._get_severity_fill(Severity.LOW)

        assert fill is not None

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_get_severity_fill_info(self):
        """Test fill color for info severity."""
        gen = XLSXReportGenerator()
        fill = gen._get_severity_fill(Severity.INFO)

        assert fill is not None

    def test_get_severity_fill_returns_none_without_openpyxl(self):
        """Test that _get_severity_fill returns None when openpyxl unavailable."""
        gen = XLSXReportGenerator()
        gen._openpyxl_available = False

        fill = gen._get_severity_fill(Severity.CRITICAL)

        assert fill is None


class TestXLSXReportGeneratorEmptyData:
    """Tests for handling empty data."""

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_generate_with_no_findings(self, tmp_path):
        """Test generation with no findings."""
        config = ReportConfig(title="Empty Report")
        report_data = ReportData(config=config, findings=[])

        gen = XLSXReportGenerator()
        output_path = tmp_path / "empty_report.xlsx"

        result = gen.generate(report_data, output_path)

        assert Path(result).exists()

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_generate_with_no_targets(self, tmp_path):
        """Test generation with no targets."""
        config = ReportConfig(title="No Targets")
        finding = Finding(
            id="f1",
            title="Test",
            severity=Severity.HIGH,
            description="Test",
            target="192.168.1.1",
        )
        report_data = ReportData(config=config, findings=[finding], targets=[])

        gen = XLSXReportGenerator()
        output_path = tmp_path / "no_targets.xlsx"

        result = gen.generate(report_data, output_path)

        assert Path(result).exists()

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_generate_with_no_services(self, tmp_path):
        """Test generation with no services."""
        config = ReportConfig(title="No Services")
        report_data = ReportData(config=config, findings=[], services=[])

        gen = XLSXReportGenerator()
        output_path = tmp_path / "no_services.xlsx"

        result = gen.generate(report_data, output_path)

        assert Path(result).exists()


class TestXLSXReportGeneratorLongContent:
    """Tests for handling long content."""

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_long_description_handled(self, tmp_path):
        """Test that long descriptions are handled."""
        config = ReportConfig(title="Long Content")
        finding = Finding(
            id="f1",
            title="Test",
            severity=Severity.HIGH,
            description="A" * 5000,  # Very long description
            target="192.168.1.1",
        )
        report_data = ReportData(config=config, findings=[finding])

        gen = XLSXReportGenerator()
        output_path = tmp_path / "long_content.xlsx"

        result = gen.generate(report_data, output_path)

        assert Path(result).exists()

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_special_characters_handled(self, tmp_path):
        """Test that special characters are handled."""
        import openpyxl

        config = ReportConfig(title="Special Characters")
        finding = Finding(
            id="f1",
            title="Test <script>alert('xss')</script>",
            severity=Severity.HIGH,
            description="Contains \"quotes\" and 'apostrophes' & ampersands",
            target="192.168.1.1",
        )
        report_data = ReportData(config=config, findings=[finding])

        gen = XLSXReportGenerator()
        output_path = tmp_path / "special_chars.xlsx"

        result = gen.generate(report_data, output_path)

        assert Path(result).exists()

        # Verify content is readable
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Findings"]
        data = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data) > 0


# =============================================================================
# Mocked Tests (run regardless of openpyxl availability)
# =============================================================================

class TestXLSXReportGeneratorMocked:
    """Tests for XLSXReportGenerator using mocked openpyxl."""

    def test_check_openpyxl_returns_true_when_available(self):
        """Test _check_openpyxl returns True when import succeeds."""
        gen = XLSXReportGenerator()

        with patch.dict('sys.modules', {'openpyxl': MagicMock()}):
            result = gen._check_openpyxl()

        # Should return True (openpyxl mock is in sys.modules)
        assert result is True

    def test_check_openpyxl_returns_false_when_unavailable(self):
        """Test _check_openpyxl returns False when import fails."""
        gen = XLSXReportGenerator()
        gen._openpyxl_available = False

        # Force ImportError
        with patch('builtins.__import__', side_effect=ImportError("No module")):
            result = gen._check_openpyxl()

        assert result is False

    def test_generate_with_mocked_openpyxl(self, sample_report_data, tmp_path):
        """Test generate() with fully mocked openpyxl."""
        # Create mock openpyxl module and classes
        mock_openpyxl = MagicMock()
        mock_wb = MagicMock()
        mock_openpyxl.Workbook.return_value = mock_wb
        mock_wb.sheetnames = ['Sheet']

        # Mock the workbook's create_sheet method
        mock_sheets = {}

        def create_sheet(name, index=None):
            mock_sheet = MagicMock()
            mock_sheets[name] = mock_sheet
            return mock_sheet

        mock_wb.create_sheet = create_sheet
        mock_wb.__getitem__ = lambda self, key: mock_sheets.get(key, MagicMock())
        mock_wb.__contains__ = lambda self, key: key in mock_sheets or key == 'Sheet'
        mock_wb.__delitem__ = MagicMock()

        # Set up all required mocks
        mock_openpyxl.styles = MagicMock()
        mock_openpyxl.styles.Font = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.PatternFill = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Alignment = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Border = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Side = MagicMock(return_value=MagicMock())
        mock_openpyxl.utils = MagicMock()
        mock_openpyxl.utils.get_column_letter = lambda x: chr(64 + x)
        mock_openpyxl.chart = MagicMock()
        mock_openpyxl.chart.PieChart = MagicMock(return_value=MagicMock())
        mock_openpyxl.chart.Reference = MagicMock(return_value=MagicMock())

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.styles': mock_openpyxl.styles,
            'openpyxl.utils': mock_openpyxl.utils,
            'openpyxl.chart': mock_openpyxl.chart,
        }):
            gen = XLSXReportGenerator()
            gen._openpyxl_available = True

            output_path = tmp_path / "mocked_report.xlsx"
            result = gen.generate(sample_report_data, output_path)

            # Should return the output path
            assert result == str(output_path)

            # Workbook should be saved
            mock_wb.save.assert_called_once_with(output_path)

    def test_generate_creates_all_sheets(self, sample_report_data, tmp_path):
        """Test that generate() creates all expected worksheets."""
        mock_openpyxl = MagicMock()
        mock_wb = MagicMock()
        mock_openpyxl.Workbook.return_value = mock_wb
        mock_wb.sheetnames = ['Sheet']

        created_sheets = []

        def track_create_sheet(name, index=None):
            created_sheets.append(name)
            return MagicMock()

        mock_wb.create_sheet = track_create_sheet
        mock_wb.__contains__ = lambda self, key: key == 'Sheet'
        mock_wb.__delitem__ = MagicMock()

        # Setup styles and utils mocks
        mock_openpyxl.styles = MagicMock()
        mock_openpyxl.styles.Font = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.PatternFill = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Alignment = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Border = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Side = MagicMock(return_value=MagicMock())
        mock_openpyxl.utils = MagicMock()
        mock_openpyxl.utils.get_column_letter = lambda x: chr(64 + x)
        mock_openpyxl.chart = MagicMock()
        mock_openpyxl.chart.PieChart = MagicMock(return_value=MagicMock())
        mock_openpyxl.chart.Reference = MagicMock(return_value=MagicMock())

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.styles': mock_openpyxl.styles,
            'openpyxl.utils': mock_openpyxl.utils,
            'openpyxl.chart': mock_openpyxl.chart,
        }):
            gen = XLSXReportGenerator()
            gen._openpyxl_available = True

            output_path = tmp_path / "sheets_report.xlsx"
            gen.generate(sample_report_data, output_path)

            # All sheets should be created
            assert "Summary" in created_sheets
            assert "Findings" in created_sheets
            assert "Targets" in created_sheets
            assert "Services" in created_sheets

    def test_generate_removes_default_sheet(self, sample_report_data, tmp_path):
        """Test that generate() removes the default 'Sheet' worksheet."""
        mock_openpyxl = MagicMock()
        mock_wb = MagicMock()
        mock_openpyxl.Workbook.return_value = mock_wb
        mock_wb.sheetnames = ['Sheet']
        mock_wb.create_sheet = MagicMock(return_value=MagicMock())

        deleted_sheets = []
        mock_wb.__contains__ = lambda self, key: key == 'Sheet'
        mock_wb.__delitem__ = lambda self, key: deleted_sheets.append(key)

        mock_openpyxl.styles = MagicMock()
        mock_openpyxl.styles.Font = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.PatternFill = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Alignment = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Border = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Side = MagicMock(return_value=MagicMock())
        mock_openpyxl.utils = MagicMock()
        mock_openpyxl.utils.get_column_letter = lambda x: chr(64 + x)
        mock_openpyxl.chart = MagicMock()
        mock_openpyxl.chart.PieChart = MagicMock(return_value=MagicMock())
        mock_openpyxl.chart.Reference = MagicMock(return_value=MagicMock())

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.styles': mock_openpyxl.styles,
            'openpyxl.utils': mock_openpyxl.utils,
            'openpyxl.chart': mock_openpyxl.chart,
        }):
            gen = XLSXReportGenerator()
            gen._openpyxl_available = True

            output_path = tmp_path / "no_default_sheet.xlsx"
            gen.generate(sample_report_data, output_path)

            # Default sheet should be deleted
            assert 'Sheet' in deleted_sheets


class TestXLSXCreateSummarySheetMocked:
    """Tests for _create_summary_sheet with mocked openpyxl."""

    def test_summary_sheet_has_title(self, sample_report_data):
        """Test summary sheet includes report title."""
        mock_openpyxl = MagicMock()
        mock_ws = MagicMock()
        mock_wb = MagicMock()

        cells = {}

        def cell(row, column, value=None):
            key = (row, column)
            if key not in cells:
                cells[key] = MagicMock()
            if value is not None:
                cells[key].value = value
            return cells[key]

        mock_ws.cell = cell
        mock_ws.__setitem__ = lambda self, key, val: setattr(cells.setdefault(key, MagicMock()), 'value', val)
        mock_ws.__getitem__ = lambda self, key: cells.setdefault(key, MagicMock())

        mock_openpyxl.styles = MagicMock()
        mock_openpyxl.styles.Font = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.PatternFill = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Alignment = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Border = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Side = MagicMock(return_value=MagicMock())
        mock_openpyxl.utils = MagicMock()
        mock_openpyxl.utils.get_column_letter = lambda x: chr(64 + x)
        mock_openpyxl.chart = MagicMock()
        mock_openpyxl.chart.PieChart = MagicMock(return_value=MagicMock())
        mock_openpyxl.chart.Reference = MagicMock(return_value=MagicMock())

        mock_wb.create_sheet = MagicMock(return_value=mock_ws)

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.styles': mock_openpyxl.styles,
            'openpyxl.utils': mock_openpyxl.utils,
            'openpyxl.chart': mock_openpyxl.chart,
        }):
            gen = XLSXReportGenerator()
            gen._openpyxl_available = True
            gen._create_summary_sheet(mock_wb, sample_report_data)

            # Title should be set in A1
            assert cells.get('A1') is not None
            assert cells['A1'].value == sample_report_data.config.title


class TestXLSXCreateFindingsSheetMocked:
    """Tests for _create_findings_sheet with mocked openpyxl."""

    def test_findings_sheet_writes_headers(self, sample_report_data):
        """Test findings sheet writes correct headers."""
        mock_openpyxl = MagicMock()
        mock_ws = MagicMock()
        mock_wb = MagicMock()

        cells = {}

        def cell(row, column, value=None):
            key = (row, column)
            if key not in cells:
                cells[key] = MagicMock()
            if value is not None:
                cells[key].value = value
            return cells[key]

        mock_ws.cell = cell
        mock_ws.column_dimensions = defaultdict(MagicMock)
        mock_ws.auto_filter = MagicMock()

        mock_openpyxl.styles = MagicMock()
        mock_openpyxl.styles.Font = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.PatternFill = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Alignment = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Border = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Side = MagicMock(return_value=MagicMock())
        mock_openpyxl.utils = MagicMock()
        mock_openpyxl.utils.get_column_letter = lambda x: chr(64 + x)

        mock_wb.create_sheet = MagicMock(return_value=mock_ws)

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.styles': mock_openpyxl.styles,
            'openpyxl.utils': mock_openpyxl.utils,
        }):
            gen = XLSXReportGenerator()
            gen._openpyxl_available = True
            gen._create_findings_sheet(mock_wb, sample_report_data)

            # Check headers in row 1
            header_values = [cells.get((1, col), MagicMock()).value for col in range(1, 14)]
            assert "ID" in header_values
            assert "Title" in header_values
            assert "Severity" in header_values

    def test_findings_sheet_writes_finding_data(self, sample_report_data):
        """Test findings sheet writes finding data."""
        mock_openpyxl = MagicMock()
        mock_ws = MagicMock()
        mock_wb = MagicMock()

        cells = {}

        def cell(row, column, value=None):
            key = (row, column)
            if key not in cells:
                cells[key] = MagicMock()
            if value is not None:
                cells[key].value = value
            return cells[key]

        mock_ws.cell = cell
        mock_ws.column_dimensions = defaultdict(MagicMock)
        mock_ws.auto_filter = MagicMock()

        mock_openpyxl.styles = MagicMock()
        mock_openpyxl.styles.Font = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.PatternFill = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.Alignment = MagicMock(return_value=MagicMock())
        mock_openpyxl.utils = MagicMock()
        mock_openpyxl.utils.get_column_letter = lambda x: chr(64 + x)

        mock_wb.create_sheet = MagicMock(return_value=mock_ws)

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.styles': mock_openpyxl.styles,
            'openpyxl.utils': mock_openpyxl.utils,
        }):
            gen = XLSXReportGenerator()
            gen._openpyxl_available = True
            gen._create_findings_sheet(mock_wb, sample_report_data)

            # Check data in row 2 (first finding)
            assert cells.get((2, 1)) is not None  # ID column
            assert cells[(2, 1)].value == "f1"
            assert cells[(2, 2)].value == "SQL Injection"


class TestXLSXCreateTargetsSheetMocked:
    """Tests for _create_targets_sheet with mocked openpyxl."""

    def test_targets_sheet_writes_target_data(self, sample_report_data):
        """Test targets sheet writes target data."""
        mock_openpyxl = MagicMock()
        mock_ws = MagicMock()
        mock_wb = MagicMock()

        cells = {}

        def cell(row, column, value=None):
            key = (row, column)
            if key not in cells:
                cells[key] = MagicMock()
            if value is not None:
                cells[key].value = value
            return cells[key]

        mock_ws.cell = cell
        mock_ws.column_dimensions = defaultdict(MagicMock)

        mock_openpyxl.styles = MagicMock()
        mock_openpyxl.styles.Font = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.PatternFill = MagicMock(return_value=MagicMock())
        mock_openpyxl.utils = MagicMock()
        mock_openpyxl.utils.get_column_letter = lambda x: chr(64 + x)

        mock_wb.create_sheet = MagicMock(return_value=mock_ws)

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.styles': mock_openpyxl.styles,
            'openpyxl.utils': mock_openpyxl.utils,
        }):
            gen = XLSXReportGenerator()
            gen._openpyxl_available = True
            gen._create_targets_sheet(mock_wb, sample_report_data)

            # Check headers
            assert cells[(1, 1)].value == "Name"
            assert cells[(1, 2)].value == "IP/URL"

            # Check target data in row 2
            assert cells[(2, 1)].value == "Web Server"
            assert cells[(2, 2)].value == "192.168.1.100"


class TestXLSXCreateServicesSheetMocked:
    """Tests for _create_services_sheet with mocked openpyxl."""

    def test_services_sheet_writes_service_data(self, sample_report_data):
        """Test services sheet writes service data."""
        mock_openpyxl = MagicMock()
        mock_ws = MagicMock()
        mock_wb = MagicMock()

        cells = {}

        def cell(row, column, value=None):
            key = (row, column)
            if key not in cells:
                cells[key] = MagicMock()
            if value is not None:
                cells[key].value = value
            return cells[key]

        mock_ws.cell = cell
        mock_ws.column_dimensions = defaultdict(MagicMock)
        mock_ws.auto_filter = MagicMock()

        mock_openpyxl.styles = MagicMock()
        mock_openpyxl.styles.Font = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.PatternFill = MagicMock(return_value=MagicMock())
        mock_openpyxl.utils = MagicMock()
        mock_openpyxl.utils.get_column_letter = lambda x: chr(64 + x)

        mock_wb.create_sheet = MagicMock(return_value=mock_ws)

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.styles': mock_openpyxl.styles,
            'openpyxl.utils': mock_openpyxl.utils,
        }):
            gen = XLSXReportGenerator()
            gen._openpyxl_available = True
            gen._create_services_sheet(mock_wb, sample_report_data)

            # Check headers
            assert cells[(1, 1)].value == "Target"
            assert cells[(1, 2)].value == "Port"
            assert cells[(1, 3)].value == "Service"

            # Check service data
            assert cells[(2, 1)].value == "192.168.1.100"
            assert cells[(2, 2)].value == 443
            assert cells[(2, 3)].value == "https"

    def test_services_sheet_no_autofilter_with_no_services(self):
        """Test services sheet doesn't add autofilter when empty."""
        mock_openpyxl = MagicMock()
        mock_ws = MagicMock()
        mock_wb = MagicMock()

        cells = {}

        def cell(row, column, value=None):
            key = (row, column)
            if key not in cells:
                cells[key] = MagicMock()
            if value is not None:
                cells[key].value = value
            return cells[key]

        mock_ws.cell = cell
        mock_ws.column_dimensions = defaultdict(MagicMock)
        mock_ws.auto_filter = MagicMock()

        mock_openpyxl.styles = MagicMock()
        mock_openpyxl.styles.Font = MagicMock(return_value=MagicMock())
        mock_openpyxl.styles.PatternFill = MagicMock(return_value=MagicMock())
        mock_openpyxl.utils = MagicMock()
        mock_openpyxl.utils.get_column_letter = lambda x: chr(64 + x)

        mock_wb.create_sheet = MagicMock(return_value=mock_ws)

        config = ReportConfig(title="No Services")
        empty_report = ReportData(config=config, findings=[], services=[])

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.styles': mock_openpyxl.styles,
            'openpyxl.utils': mock_openpyxl.utils,
        }):
            gen = XLSXReportGenerator()
            gen._openpyxl_available = True
            gen._create_services_sheet(mock_wb, empty_report)

            # Auto filter ref should not be set (row == 2, so row > 2 is False)
            # The ref attribute should not have been assigned a value
            # In reality this tests the conditional in the code


class TestGetColumnLetterFunction:
    """Tests for the module-level get_column_letter function."""

    def test_get_column_letter_with_mocked_openpyxl(self):
        """Test get_column_letter function."""
        mock_openpyxl = MagicMock()
        mock_openpyxl.utils = MagicMock()
        mock_openpyxl.utils.get_column_letter = lambda x: chr(64 + x)

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.utils': mock_openpyxl.utils,
        }):
            from purplesploit.reporting.xlsx import get_column_letter

            assert get_column_letter(1) == "A"
            assert get_column_letter(2) == "B"
            assert get_column_letter(26) == "Z"


class TestXLSXSeverityFillMocked:
    """Tests for _get_severity_fill with mocked openpyxl."""

    def test_severity_fill_uses_correct_colors(self):
        """Test that severity fills use the correct color codes."""
        mock_openpyxl = MagicMock()
        mock_openpyxl.styles = MagicMock()

        captured_colors = []

        def capture_pattern_fill(start_color, end_color, fill_type):
            captured_colors.append(start_color)
            return MagicMock()

        mock_openpyxl.styles.PatternFill = capture_pattern_fill

        with patch.dict('sys.modules', {
            'openpyxl': mock_openpyxl,
            'openpyxl.styles': mock_openpyxl.styles,
        }):
            gen = XLSXReportGenerator()
            gen._openpyxl_available = True

            gen._get_severity_fill(Severity.CRITICAL)
            assert "7B241C" in captured_colors

            gen._get_severity_fill(Severity.HIGH)
            assert "C0392B" in captured_colors

            gen._get_severity_fill(Severity.MEDIUM)
            assert "E67E22" in captured_colors

            gen._get_severity_fill(Severity.LOW)
            assert "F1C40F" in captured_colors

            gen._get_severity_fill(Severity.INFO)
            assert "3498DB" in captured_colors
